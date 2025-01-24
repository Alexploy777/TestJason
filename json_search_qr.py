import json
import re

import openpyxl
from openpyxl.styles import NamedStyle


class JsonSearch:
    def __init__(self, json_file_name, sign_of_qr, sign_for_search_list):
        self.json_data = None
        self.json_reading(json_file_name)
        self.sign_of_qr = sign_of_qr
        self.sign_status = sign_for_search_list[0]
        self.sign_erors = sign_for_search_list[1]

    def json_reading(self, json_file_name):
        try:
            with open(json_file_name, 'r', encoding='utf-8') as f:
                text_data = f.read()
                text_data = text_data.replace('][', ',')
                json_data = json.loads(text_data)
            self.json_data = json_data
        except Exception as e:
            print(f'Произошла ошибка {e} чтения или обработки файла')


    def decoder(self, string):
        return string.encode('latin1').decode('utf-8')

    def get_status(self):
        qr_data_dic = {}
        for block in self.json_data:
            dic_of_block = block["cisInfo"]
            requestedCis = dic_of_block[self.sign_of_qr]
            if self.sign_status in dic_of_block:
                qr_data_dic[requestedCis] = dic_of_block[self.sign_status]
            else:
                qr_data_dic[requestedCis] = self.decoder(block[self.sign_erors])
        self.qr_data_dic = qr_data_dic
        self.write_to_file()
        self.write_to_excel()
        print('Обработка завершена!')

    def write_to_file(self):
        with open('qr_data_result.json', 'w', encoding='utf-8') as f:
            json.dump(self.qr_data_dic, f, ensure_ascii=False, indent=0)

    def write_to_excel(self):
        # Создаем новую книгу Excel
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "QR_check"

        def sanitize_string(value):
            if isinstance(value, str):
                # Удаляем все символы, не входящие в допустимый диапазон
                return re.sub(r"[\x00-\x1F\x7F-\x9F]", "", value)
            return value

        # Определяем стиль для текста
        text_style = NamedStyle(name="text_style", number_format="@")
        wb.add_named_style(text_style)

        sheet.cell(row=1, column=1, value="QR").style = text_style  # Заголовок первого столбца
        sheet.cell(row=1, column=2, value="Статус").style = text_style  # Заголовок второго столбца

        for row, (key, value) in enumerate(self.qr_data_dic.items(), start=2):
            sanitized_key = sanitize_string(key)  # Очищаем ключи
            sanitized_value = sanitize_string(value)  # Очищаем значения
            cell_key = sheet.cell(row=row, column=1, value=sanitized_key)
            cell_value = sheet.cell(row=row, column=2, value=sanitized_value)
            cell_key.style = text_style
            cell_value.style = text_style

        # Сохраняем в файл
        wb.save("qr_check_excel.xlsx")
        print("Словарь успешно записан в файл qr_check_excel.xlsx")




if __name__ == '__main__':
    input_file_path = 'json_data/check_short.json'  # файл с json
    # input_file_path = 'json_data/answer_mini.json'  # файл с json
    sign_of_qr = 'cis'  # ключ индентификатора QR
    sign_for_search_list = ["status", "errorMessage"]
    jsonsearch = JsonSearch(input_file_path, sign_of_qr, sign_for_search_list)
    result = jsonsearch.get_status()
